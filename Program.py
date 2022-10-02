import openpyxl
import random
book = openpyxl.load_workbook("Task 1.xlsx", read_only=True)
sheetname = book.worksheets[0]
sheet = book.worksheets[1]
sheet1= book.worksheets[2]
sheet3= book.worksheets[3]


#--------------------------------------------------------- First
variants = list()
for row in range(3, sheet.max_row + 1):
    variants.append(sheet[row][0].value)

t1y1 = list()
for row in range(3, sheet.max_row + 1):
    t1y1.append(sheet[row][1].value)

t1m1 = list()
for row in range(3, sheet.max_row + 1):
    t1m1.append(sheet[row][2].value)

t1y2 = list()
for row in range(3, sheet.max_row + 1):
    t1y2.append(sheet[row][3].value)

t1m2 = list()
for row in range(3, sheet.max_row + 1):
    t1m2.append(sheet[row][4].value)

t1v1 = list()
for row in range(3, sheet.max_row + 1):
    t1v1.append(sheet[row][5].value)

t1Tmoor = list()
for row in range(3, sheet.max_row + 1):
    t1Tmoor.append(sheet[row][6].value)

# Delta T
dt = list()
for i in range(len(variants)):
    dtr = (t1y2[i] - t1y1[i]) * 12 + t1m2[i] - t1m1[i]
    dt.append(dtr)

# V2 memory
t1v2 = list()
for i in range(len(variants)):
    t1v2r = t1v1[i] * 2**(dt[i]/t1Tmoor[i])
    t1v2r = round(t1v2r, 2)
    t1v2.append(t1v2r)

#-------------------------------------------- Second

t2y1 = list()
for row in range(3, sheet1.max_row + 1):
    t2y1.append(sheet1[row][1].value)

t2m1 = list()
for row in range(3, sheet1.max_row + 1):
    t2m1.append(sheet1[row][2].value)

t2y2 = list()
for row in range(3, sheet1.max_row + 1):
    t2y2.append(sheet1[row][3].value)

t2m2 = list()
for row in range(3, sheet1.max_row + 1):
    t2m2.append(sheet1[row][4].value)

t2v1 = list()
for row in range(3, sheet1.max_row + 1):
    t2v1.append(sheet1[row][5].value)

t2Tmoor = list()
for row in range(3, sheet1.max_row + 1):
    t2Tmoor.append(sheet1[row][6].value)
#-------------------------------------------

cells = sheet1['B3':'G34']
mouth__dif = 0
result_=list()
for Y1, M1, Y2, M2, V2, Tmoor in cells:
    mouth__dif = ((Y2.value - Y1.value) * 12 + M2.value - M1.value) / Tmoor.value
    # print(mouth__dif)

    power = pow(mouth__dif , 2 )
    # print (power)

    result = V2.value/power
    result = round(result , 2)
    result_.append(result)

# ------------------------------ Third

t3y1 = list()
for row in range(3, sheet3.max_row):
    t3y1.append(sheet3[row][1].value)

t3m1 = list()
for row in range(3, sheet3.max_row):
    t3m1.append(sheet3[row][2].value)

t3y2 = list()
for row in range(3, sheet3.max_row):
    t3y2.append(sheet3[row][3].value)

t3m2 = list()
for row in range(3, sheet3.max_row):
    t3m2.append(sheet3[row][4].value)

t3v1 = list()
for row in range(3, sheet3.max_row):
    t3v1.append(sheet3[row][5].value)

t3v2 = list()
for row in range(3, sheet3.max_row):
    t3v2.append(sheet3[row][6].value)

# --------------------------------------------

book = list()             			#lists for containing data
variants_list = list()
answers = list()						#list with answers for teacher
t = True
n = 3

def make_variant(n):   #function for reading the data for each variant
	book.clear()
	for cell in range(0, 7):
		personal_variant = sheet3[n][cell].value
		book.append(personal_variant)

		'''[0, 2011, 1, 2017, 2, 512, 65536, ]
		    ^	 ^   ^    ^   ^   ^     ^
		    V    Y1  M1   Y2  M2  V1    V2
		'''

def calculate_and_write_variant(variant):
	v = variant[0]
	y1 = variant[1]
	m1 = variant[2]
	y2 = variant[3]
	m2 = variant[4]
	v1 = variant[5]
	v2 = variant[6]

	t_ = (y2 - y1) * 12 + m2 - m1
	a = v2 / v1 #how many times increased
	T = 2 * t_ / a
	T = round(T, 2) #round to 2 numbers after points
	answers.append(T)

while t:
	make_variant(n)
	calculate_and_write_variant(book)
	#space to function, which make txt file with task
	#space to function, which generate the answer and make txt file
	n = n + 1
	if n == 35:
		t = False


#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#------------------------------------- Start main-------------------------------
#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

c = list() #Name list ----------------------------------------------------------
for row in range(2, sheetname.max_row + 1):
    c.append(sheetname[row][1].value)
random.shuffle(c)
a = list(filter(None, c))


#answer1 = t1v2 # Answer to task 1, 2, 3 in order to search len variants --------------------------------------
b = list()
b.append(t1v2)
b.append(result_)
b.append(answers)


if len(b) < len(a): # Start logic expression ----------------------------------------
    Max = a
    Min = b
else: Max = b; Min = a
LenMax = len(Max)

for i in range(LenMax): # Start write to txt file
    if a == Max:
        bb = (i % len(Min))
        aa = i
    else:
        aa = (i % len(Min))
        bb = i
    name = "ЛР2_Закон Мура " + str(a[aa])
    my_file = '{}.txt'.format(name)
    with open(my_file, "w") as f:
        f.write("- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   \n")
        f.write("ОМ ЛР 02. Закон Мура. Варіант" + "(" + str(i+1) + ")\n")
        f.write("Завдання 1. ------------------------- \n")
        f.write("Дано: \n")
        f.write(" Y1   M1  Y2   M2  V1  Tmoor \n")
        f.write(str(t1y1[bb]) +"  "+str(t1m1[bb]) +"  "+str(t1y2[bb]) +"  "+str(t1m2[bb]) +"  "+str(t1v1[bb]) +"  "+str(t1Tmoor[bb]) +"  "  + "\n \v")
        f.write(
                "В M1 місяці Y1 року флеш пристрій ємністю V1 коштував 70$.\n"
                "1. Знайти ємність V2, Мб, флеш пристрію такої же користувацької цінності\n"
                "в Y2 році M2 місяці.\n"
                "Вважаємо, що флеш-накопичувачі подвоюють свою ємність кожні Tmoor місяців.\n \v")
        f.write("Завдання 2. -------------------------\n")
        f.write("Дано: \n")
        f.write(" Y1   M1  Y2   M2  V1  Tmoor \n")
        f.write(str(t2y1[bb]) + "  " + str(t2m1[bb]) + "  " + str(t2y2[bb]) + "  " + str(t2m2[bb]) + "  " + str(
            t2v1[bb]) + "  " + str(t2Tmoor[bb]) + "  " + "\n \v")
        f.write("В M1 місяці Y1 року флеш пристрій коштував 70$. Курс долара 5 грн.\n"
                "1. Знайти його ємність V1, Мб, якщо \n"
                "в Y2 році M2 місяці флеш пристрій ємністю V2 коштував 400 грн. Курс долара 26 грн.\n"
                "Вважаємо, що флеш-накопичувачі подвоюють свою ємність кожні Tmoor місяців." "\n \v")
        f.write("Завдання 3. -------------------------\n")
        f.write("Дано: \n")
        f.write(" Y1   M1  Y2   M2  V1  V2 \n")
        f.write(str(t3y1[bb]) +"  "+str(t3m1[bb]) +"  "+str(t3y2[bb]) +"  "+str(t3m2[bb]) +"  "+str(t3v1[bb]) +"  "+str(t3v2[bb]) +"  "  + "\n \v")
        f.write("В M1 місяці Y1 року флеш пристрій коштував 70$. Курс долара 5 грн.\n"
                "в Y2 році M2 місяці флеш пристрій ємністю V2 коштував 400 грн. Курс долара 26 грн.\n"
                "Знайти період Tmoor (місяців), за якій показники V пристрою збільшуються вдвічі\n")

my_answer = '1Answer.txt'
for i in range(LenMax): # Write Answer -----------------------
    if a == Max:
        bb = (i % len(Min))
        aa = i
    else:
        aa = (i % len(Min))
        bb = i
    with open(my_answer, "a") as ma:
        ma.write(str(a[aa]) + " Answer Laba is: \n"  
        "Task 1 --- " + str(t1v2[bb]) + "\n"
        "Task 2 --- " + str(result_[bb]) + "\n"
        "Task 3 --- " + str(answers[bb]) + "\v")











